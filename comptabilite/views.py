import csv
import io

from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import dateformat
from django.utils.translation import get_language

from users.mixins import role_required

from .models import Budget, DepenseMensuelle, RecetteMensuelle

MOIS_AR_MAROC = {
    1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل", 5: "ماي", 6: "يونيو",
    7: "يوليوز", 8: "غشت", 9: "شتنبر", 10: "أكتوبر", 11: "نونبر", 12: "دجنبر",
}


def _mois_label(mois):
    if get_language() == "ar":
        return f"{MOIS_AR_MAROC[mois.month]} {mois.year}"
    return dateformat.format(mois, "F Y")


def _repartition(queryset, group_field, choices, total):
    labels = dict(choices)
    lignes = list(
        queryset.values(group_field)
        .annotate(montant=Sum("montant"))
        .order_by("-montant")
    )
    for ligne in lignes:
        cle = ligne[group_field]
        ligne["label"] = labels.get(cle, cle)
        ligne["montant"] = ligne["montant"] or 0
        ligne["pourcentage"] = round((ligne["montant"] / total) * 100) if total else 0
    return lignes


def _table_mensuelle(programmes):
    mois_qs = DepenseMensuelle.objects.order_by("mois").values_list("mois", flat=True).distinct()
    lignes_brutes = {
        (d.mois, d.programme): d.montant for d in DepenseMensuelle.objects.all()
    }
    lignes = []
    for mois in mois_qs:
        valeurs = [lignes_brutes.get((mois, code), 0) for code, _label in programmes]
        lignes.append(
            {"mois": mois, "mois_label": _mois_label(mois), "valeurs": valeurs, "total": sum(valeurs)}
        )
    return lignes


def _donnees_consolidation():
    total_depenses = DepenseMensuelle.objects.aggregate(total=Sum("montant"))["total"] or 0
    total_recettes = RecetteMensuelle.objects.aggregate(total=Sum("montant"))["total"] or 0

    depenses_par_programme = _repartition(
        DepenseMensuelle.objects.all(), "programme", DepenseMensuelle.Programme.choices, total_depenses
    )
    recettes_par_source = _repartition(
        RecetteMensuelle.objects.all(), "source", RecetteMensuelle.Source.choices, total_recettes
    )

    return {
        "total_depenses": total_depenses,
        "total_recettes": total_recettes,
        "resultat": total_recettes - total_depenses,
        "depenses_par_programme": depenses_par_programme,
        "recettes_par_source": recettes_par_source,
        "top_depenses": depenses_par_programme[:3],
        "table_mensuelle": _table_mensuelle(DepenseMensuelle.Programme.choices),
        "programmes": DepenseMensuelle.Programme.choices,
    }


@role_required("responsable")
def consolidation(request):
    return render(request, "comptabilite/consolidation.html", _donnees_consolidation())


@role_required("responsable")
def bilan_financier(request):
    periode = request.GET.get("periode", "")
    budgets = Budget.objects.all().order_by("-periode", "module")
    if periode:
        budgets = budgets.filter(periode=periode)
    periodes = (
        Budget.objects.order_by("-periode").values_list("periode", flat=True).distinct()
    )
    totaux = budgets.aggregate(total_recettes=Sum("recettes"), total_depenses=Sum("depenses"))
    total_recettes = totaux["total_recettes"] or 0
    total_depenses = totaux["total_depenses"] or 0
    return render(
        request,
        "comptabilite/bilan_financier.html",
        {
            "budgets": budgets,
            "periode": periode,
            "periodes": periodes,
            "total_recettes": total_recettes,
            "total_depenses": total_depenses,
            "total_solde": total_recettes - total_depenses,
        },
    )


@role_required("responsable")
def exporter_donnees(request):
    periode = request.GET.get("periode", "")
    budgets = Budget.objects.all().order_by("periode", "module")
    if periode:
        budgets = budgets.filter(periode=periode)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="bilan_comptable.csv"'
    writer = csv.writer(response)
    writer.writerow(["Module", "Période", "Recettes", "Dépenses", "Solde"])
    for budget in budgets:
        writer.writerow(
            [
                budget.get_module_display(),
                budget.periode,
                budget.recettes,
                budget.depenses,
                budget.solde,
            ]
        )
    return response


@role_required("responsable")
def exporter_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    donnees = _donnees_consolidation()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Dépenses par programme"
    ws1.append(["Programme", "Montant (MAD)", "%"])
    for ligne in donnees["depenses_par_programme"]:
        ws1.append([str(ligne["label"]), float(ligne["montant"]), ligne["pourcentage"]])
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("Recettes par source")
    ws2.append(["Source", "Montant (MAD)", "%"])
    for ligne in donnees["recettes_par_source"]:
        ws2.append([str(ligne["label"]), float(ligne["montant"]), ligne["pourcentage"]])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    ws3 = wb.create_sheet("Budget mensuel")
    labels = [str(label) for _code, label in donnees["programmes"]]
    ws3.append(["Mois", *labels, "Total"])
    for ligne in donnees["table_mensuelle"]:
        ws3.append([ligne["mois"].strftime("%Y-%m"), *[float(v) for v in ligne["valeurs"]], float(ligne["total"])])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    ws4 = wb.create_sheet("Résumé")
    ws4.append(["Total des ressources (MAD)", float(donnees["total_recettes"])])
    ws4.append(["Total des dépenses (MAD)", float(donnees["total_depenses"])])
    ws4.append(["Résultat (MAD)", float(donnees["resultat"])])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bilan_comptable.xlsx"'
    return response


@role_required("responsable")
def exporter_pdf(request):
    from django.utils import translation
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    # reportlab's built-in fonts can't render Arabic glyphs, so force French
    # labels for this export regardless of the active site language. The
    # lazy translation proxies must be resolved to plain str while the
    # override is still active, or they'd fall back to the request's
    # actual active language once this block exits.
    with translation.override("fr"):
        donnees = _donnees_consolidation()
        for ligne in donnees["depenses_par_programme"]:
            ligne["label"] = str(ligne["label"])
        for ligne in donnees["recettes_par_source"]:
            ligne["label"] = str(ligne["label"])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Association Bassma Maroc — Bilan financier", styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            f"Total des ressources : {donnees['total_recettes']} MAD &nbsp;&nbsp; "
            f"Total des dépenses : {donnees['total_depenses']} MAD &nbsp;&nbsp; "
            f"Résultat : {donnees['resultat']} MAD",
            styles["Normal"],
        ),
        Spacer(1, 16),
        Paragraph("Dépenses par programme", styles["Heading2"]),
    ]

    data = [["Programme", "Montant (MAD)", "%"]]
    for ligne in donnees["depenses_par_programme"]:
        data.append([str(ligne["label"]), f"{ligne['montant']:,}", f"{ligne['pourcentage']}%"])
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14432B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph("Recettes par source", styles["Heading2"]))

    data2 = [["Source", "Montant (MAD)", "%"]]
    for ligne in donnees["recettes_par_source"]:
        data2.append([str(ligne["label"]), f"{ligne['montant']:,}", f"{ligne['pourcentage']}%"])
    table2 = Table(data2, hAlign="LEFT")
    table2.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14432B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements.append(table2)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="bilan_comptable.pdf"'
    return response
